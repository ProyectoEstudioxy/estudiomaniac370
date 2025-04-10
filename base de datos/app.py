rom flask import Flask, render_template, request, redirect, url_for
import sqlite3
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
from openpyxl import Workbook

BASE_DIR = r"C:\Users\Guillermo\OneDrive\Documentos\AAA PYTHON\ESTUDIO MANIAC 370"
app = Flask(__name__, 
            template_folder=os.path.join(BASE_DIR, 'templates'), 
            static_folder=os.path.join(BASE_DIR, 'static'))
DB_PATH = os.path.join(BASE_DIR, 'database.db')

# Crear o actualizar la tabla
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # Crear una nueva tabla con los campos adicionales
    c.execute('''CREATE TABLE IF NOT EXISTS registros (
                 id INTEGER PRIMARY KEY AUTOINCREMENT, 
                 fecha TEXT, 
                 cuenta TEXT, 
                 descripcion TEXT, 
                 saldo_inicio REAL, 
                 debe REAL, 
                 haber REAL)''')
    conn.commit()
    conn.close()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/formulario')
def formulario():
    return render_template('formulario.html')

@app.route('/guardar', methods=['POST'])
def guardar():
    fecha = request.form['fecha']
    cuenta = request.form['cuenta']
    descripcion = request.form['descripcion']
    saldo_inicio = float(request.form['saldo_inicio'] or 0)
    debe = float(request.form['debe'] or 0)
    haber = float(request.form['haber'] or 0)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO registros (fecha, cuenta, descripcion, saldo_inicio, debe, haber) VALUES (?, ?, ?, ?, ?, ?)", 
              (fecha, cuenta, descripcion, saldo_inicio, debe, haber))
    conn.commit()
    conn.close()
    return redirect(url_for('consulta'))

@app.route('/consulta')
def consulta():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM registros")
    registros = c.fetchall()
    conn.close()
    return render_template('consulta.html', registros=registros)

@app.route('/exportar_excel')
def exportar_excel():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM registros")
    registros = c.fetchall()
    conn.close()

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance"

    # Encabezados de la hoja de transacciones
    ws['A1'] = "ID"
    ws['B1'] = "Fecha"
    ws['C1'] = "Cuenta"
    ws['D1'] = "Descripción"
    ws['E1'] = "Saldo Inicio"
    ws['F1'] = "Debe"
    ws['G1'] = "Haber"
    ws['H1'] = "Saldo Cierre"

    # Llenar transacciones
    for i, registro in enumerate(registros, start=2):
        ws[f'A{i}'] = registro[0]  # ID
        ws[f'B{i}'] = registro[1]  # Fecha
        ws[f'C{i}'] = registro[2]  # Cuenta
        ws[f'D{i}'] = registro[3]  # Descripción
        ws[f'E{i}'] = registro[4]  # Saldo Inicio
        ws[f'F{i}'] = registro[5]  # Debe
        ws[f'G{i}'] = registro[6]  # Haber
        # Fórmula para Saldo Cierre: Saldo Inicio + Debe - Haber
        ws[f'H{i}'] = f"=E{i}+F{i}-G{i}"

    # Crear hoja de mayores
    ws_mayores = wb.create_sheet("Mayores")
    ws_mayores['A1'] = "Cuenta"
    ws_mayores['B1'] = "Saldo Inicio"
    ws_mayores['C1'] = "Total Debe"
    ws_mayores['D1'] = "Total Haber"
    ws_mayores['E1'] = "Saldo Cierre"

    # Obtener cuentas únicas y calcular mayores
    cuentas = set([r[2] for r in registros])  # Lista de cuentas únicas
    for i, cuenta in enumerate(cuentas, start=2):
        ws_mayores[f'A{i}'] = cuenta
        # Sumar Saldo Inicio, Debe y Haber para esta cuenta
        ws_mayores[f'B{i}'] = sum(r[4] for r in registros if r[2] == cuenta)
        ws_mayores[f'C{i}'] = sum(r[5] for r in registros if r[2] == cuenta)
        ws_mayores[f'D{i}'] = sum(r[6] for r in registros if r[2] == cuenta)
        # Saldo Cierre = Saldo Inicio + Debe - Haber
        ws_mayores[f'E{i}'] = f"=B{i}+C{i}-D{i}"

    # Guardar el archivo
    excel_path = os.path.join(BASE_DIR, 'balance.xlsx')
    wb.save(excel_path)
    return f"Excel generado en {excel_path}"

if __name__ == '__main__':
    init_db()
    app.run(debug=True)