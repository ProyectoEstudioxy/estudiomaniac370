from flask import Flask, render_template, request, redirect, url_for
import sqlite3
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
from openpyxl import Workbook

# Ruta base temporal fuera de OneDrive
BASE_DIR = r"C:\Temp"
app = Flask(__name__, 
            template_folder=os.path.join(BASE_DIR, 'templates'), 
            static_folder=os.path.join(BASE_DIR, 'static'))
DB_PATH = os.path.join(BASE_DIR, 'database.db')

# Inicializar la base de datos
def init_db():
    try:
        # Asegurarse de que la carpeta exista
        os.makedirs(BASE_DIR, exist_ok=True)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS registros (
                     id INTEGER PRIMARY KEY AUTOINCREMENT, 
                     fecha TEXT, 
                     cuenta TEXT, 
                     descripcion TEXT, 
                     saldo_inicio REAL, 
                     debe REAL, 
                     haber REAL)''')
        conn.commit()
        conn.close()
        print(f"Base de datos creada o abierta correctamente en {DB_PATH}")
    except sqlite3.OperationalError as e:
        print(f"Error al abrir la base de datos: {e}")
        raise

# Resto del código (sin cambios)
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/formulario')
def formulario():
    return render_template('formulario.html')

@app.route('/guardar', methods=['POST'])
def guardar():
    fecha = request.form['fecha']
    cuenta = request.form['cuenta']
    descripcion = request.form['descripcion']
    saldo_inicio = float(request.form.get('saldo_inicio', 0))
    debe = float(request.form.get('debe', 0))
    haber = float(request.form.get('haber', 0))
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO registros (fecha, cuenta, descripcion, saldo_inicio, debe, haber) VALUES (?, ?, ?, ?, ?, ?)", 
              (fecha, cuenta, descripcion, saldo_inicio, debe, haber))
    conn.commit()
    conn.close()
    return redirect(url_for('consulta'))

@app.route('/consulta')
def consulta():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM registros")
    registros = c.fetchall()
    conn.close()
    return render_template('consulta.html', registros=registros)

@app.route('/exportar_excel')
def exportar_excel():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM registros")
    registros = c.fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance"
    ws['A1'] = "ID"
    ws['B1'] = "Fecha"
    ws['C1'] = "Cuenta"
    ws['D1'] = "Descripción"
    ws['E1'] = "Saldo Inicio"
    ws['F1'] = "Debe"
    ws['G1'] = "Haber"
    ws['H1'] = "Saldo Cierre"
    for i, registro in enumerate(registros, start=2):
        ws[f'A{i}'] = registro[0]
        ws[f'B{i}'] = registro[1]
        ws[f'C{i}'] = registro[2]
        ws[f'D{i}'] = registro[3]
        ws[f'E{i}'] = registro[4]
        ws[f'F{i}'] = registro[5]
        ws[f'G{i}'] = registro[6]
        ws[f'H{i}'] = f"=E{i}+F{i}-G{i}"
    ws_mayores = wb.create_sheet("Mayores")
    ws_mayores['A1'] = "Cuenta"
    ws_mayores['B1'] = "Saldo Inicio"
    ws_mayores['C1'] = "Total Debe"
    ws_mayores['D1'] = "Total Haber"
    ws_mayores['E1'] = "Saldo Cierre"
    cuentas = set([r[2] for r in registros])
    for i, cuenta in enumerate(cuentas, start=2):
        ws_mayores[f'A{i}'] = cuenta
        ws_mayores[f'B{i}'] = sum(r[4] for r in registros if r[2] == cuenta)
        ws_mayores[f'C{i}'] = sum(r[5] for r in registros if r[2] == cuenta)
        ws_mayores[f'D{i}'] = sum(r[6] for r in registros if r[2] == cuenta)
        ws_mayores[f'E{i}'] = f"=B{i}+C{i}-D{i}"
    excel_path = os.path.join(BASE_DIR, 'balance.xlsx')
    wb.save(excel_path)
    return f"Excel generado en {excel_path}"

if __name__ == '__main__':
    init_db()
    app.run(debug=True)