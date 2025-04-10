from flask import Flask, render_template, request, redirect, url_for
import sqlite3
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Configuración
BASE_DIR = r"C:\Users\Guillermo\OneDrive\Documentos\AAA PYTHON\ESTUDIO MANIAC 370"
app = Flask(__name__, 
            template_folder=os.path.join(BASE_DIR, 'templates'), 
            static_folder=os.path.join(BASE_DIR, 'static'))
DB_PATH = os.path.join(BASE_DIR, 'formularios.db')
EXCEL_PATH = os.path.join(BASE_DIR, 'formularios.xlsx')

# Inicializar la base de datos
def init_db():
    os.makedirs(BASE_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS formularios (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 numero_formulario INTEGER,
                 fecha_emision TEXT,
                 empresa_emite TEXT,
                 tipo_orden TEXT,
                 origen_empresa TEXT,
                 origen_cuenta TEXT,
                 origen_detalle TEXT,
                 origen_monto REAL,
                 origen_observacion TEXT,
                 origen_mail TEXT,
                 origen_whatsapp TEXT,
                 origen_cargar INTEGER,
                 origen_tipo_destino TEXT,
                 destino_empresa TEXT,
                 destino_cuenta TEXT,
                 destino_detalle TEXT,
                 destino_monto REAL,
                 destino_observacion TEXT,
                 destino_mail TEXT,
                 destino_whatsapp TEXT,
                 destino_cargar INTEGER,
                 destino_tipo_destino TEXT,
                 doc_factura INTEGER,
                 doc_recibo INTEGER,
                 doc_fotocopia_cma INTEGER,
                 doc_presupuesto INTEGER,
                 doc_ticket_transferencia INTEGER,
                 doc_ticket_otros INTEGER,
                 doc_constancia_cbu INTEGER,
                 doc_constancia_cuit INTEGER,
                 doc_otros INTEGER,
                 doc_otros_detalle TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS cuenta_corriente (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 formulario_id INTEGER,
                 fecha TEXT,
                 detalle TEXT,
                 saldo_inicio REAL,
                 debe REAL,
                 haber REAL,
                 saldo_cierre REAL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS forma_pago (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 formulario_id INTEGER,
                 tipo TEXT,
                 detalle TEXT,
                 monto REAL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS forma_cobro (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 formulario_id INTEGER,
                 tipo TEXT,
                 detalle TEXT,
                 monto REAL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS contador (
                 id INTEGER PRIMARY KEY,
                 numero INTEGER)''')
    c.execute("INSERT OR IGNORE INTO contador (id, numero) VALUES (1, 0)")
    conn.commit()
    conn.close()

# Obtener el número correlativo
def get_next_numero():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT numero FROM contador WHERE id = 1")
    numero = c.fetchone()[0]
    c.execute("UPDATE contador SET numero = numero + 1 WHERE id = 1")
    conn.commit()
    conn.close()
    return numero + 1

# Guardar en Excel
def save_to_excel(formulario_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM formularios WHERE id = ?", (formulario_id,))
    formulario = c.fetchone()
    c.execute("SELECT * FROM cuenta_corriente WHERE formulario_id = ?", (formulario_id,))
    cuenta_corriente = c.fetchall()
    c.execute("SELECT * FROM forma_pago WHERE formulario_id = ?", (formulario_id,))
    forma_pago = c.fetchall()
    c.execute("SELECT * FROM forma_cobro WHERE formulario_id = ?", (formulario_id,))
    forma_cobro = c.fetchall()
    conn.close()

    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

    ws = wb.create_sheet(f"Formulario_{formulario[1]}")
    ws.append(["Número de Formulario", formulario[1]])
    ws.append(["Fecha de Emisión", formulario[2]])
    ws.append(["Empresa que Emite", formulario[3]])
    ws.append(["Tipo de Orden", formulario[4]])
    ws.append([])
    ws.append(["Origen de Fondos"])
    ws.append(["Denominación de la Empresa", formulario[5]])
    ws.append(["Cuenta según el Plan", formulario[6]])
    ws.append(["Detalle", formulario[7]])
    ws.append(["Monto", formulario[8]])
    ws.append(["Observación", formulario[9]])
    ws.append(["Mail de Destino", formulario[10]])
    ws.append(["Grupo de WhatsApp", formulario[11]])
    ws.append(["Cargar en el Sistema", "Sí" if formulario[12] else "No"])
    ws.append(["Tipo de Destino de la Operación", formulario[13]])
    ws.append([])
    ws.append(["Destino de Fondos"])
    ws.append(["Denominación de la Empresa", formulario[14]])
    ws.append(["Cuenta según el Plan", formulario[15]])
    ws.append(["Detalle", formulario[16]])
    ws.append(["Monto", formulario[17]])
    ws.append(["Observación", formulario[18]])
    ws.append(["Mail de Destino", formulario[19]])
    ws.append(["Grupo de WhatsApp", formulario[20]])
    ws.append(["Cargar en el Sistema", "Sí" if formulario[21] else "No"])
    ws.append(["Tipo de Destino de la Operación", formulario[22]])
    ws.append([])
    ws.append(["Cuenta Corriente"])
    ws.append(["Fecha", "Detalle", "Saldo al Inicio", "Debe", "Haber", "Saldo al Cierre"])
    for cc in cuenta_corriente:
        ws.append([cc[2], cc[3], cc[4], cc[5], cc[6], cc[7]])
    ws.append([])
    ws.append(["Forma de Pago"])
    ws.append(["Tipo", "Detalle", "Monto"])
    for fp in forma_pago:
        ws.append([fp[2], fp[3], fp[4]])
    ws.append([])
    ws.append(["Forma de Cobro"])
    ws.append(["Tipo", "Detalle", "Monto"])
    for fc in forma_cobro:
        ws.append([fc[2], fc[3], fc[4]])
    ws.append([])
    ws.append(["Documentación Adjunta"])
    ws.append(["Factura", "Sí" if formulario[23] else "No"])
    ws.append(["Recibo", "Sí" if formulario[24] else "No"])
    ws.append(["Fotocopia CMA", "Sí" if formulario[25] else "No"])
    ws.append(["Presupuesto", "Sí" if formulario[26] else "No"])
    ws.append(["Ticket de Transferencia", "Sí" if formulario[27] else "No"])
    ws.append(["Ticket de Otros Medios", "Sí" if formulario[28] else "No"])
    ws.append(["Constancia de CBU", "Sí" if formulario[29] else "No"])
    ws.append(["Constancia de CUIT", "Sí" if formulario[30] else "No"])
    ws.append(["Otros", "Sí" if formulario[31] else "No"])
    ws.append(["Detalles de Otros", formulario[32]])

    wb.save(EXCEL_PATH)

@app.route('/')
def index():
    return render_template('formulario.html', numero_formulario=get_next_numero())

@app.route('/guardar', methods=['POST'])
def guardar():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Datos generales
    numero_formulario = get_next_numero()
    fecha_emision = request.form['fecha_emision']
    empresa_emite = request.form['empresa_emite']
    tipo_orden = request.form['tipo_orden']

    # Origen de fondos
    origen_empresa = request.form['origen_empresa']
    origen_cuenta = request.form['origen_cuenta']
    origen_detalle = request.form['origen_detalle']
    origen_monto = float(request.form['origen_monto'] or 0)
    origen_observacion = request.form['origen_observacion']
    origen_mail = request.form['origen_mail']
    origen_whatsapp = request.form['origen_whatsapp']
    origen_cargar = 1 if 'origen_cargar' in request.form else 0
    origen_tipo_destino = request.form['origen_tipo_destino']

    # Destino de fondos
    destino_empresa = request.form['destino_empresa']
    destino_cuenta = request.form['destino_cuenta']
    destino_detalle = request.form['destino_detalle']
    destino_monto = float(request.form['destino_monto'] or 0)
    destino_observacion = request.form['destino_observacion']
    destino_mail = request.form['destino_mail']
    destino_whatsapp = request.form['destino_whatsapp']
    destino_cargar = 1 if 'destino_cargar' in request.form else 0
    destino_tipo_destino = request.form['destino_tipo_destino']

    # Documentación adjunta
    doc_factura = 1 if 'doc_factura' in request.form else 0
    doc_recibo = 1 if 'doc_recibo' in request.form else 0
    doc_fotocopia_cma = 1 if 'doc_fotocopia_cma' in request.form else 0
    doc_presupuesto = 1 if 'doc_presupuesto' in request.form else 0
    doc_ticket_transferencia = 1 if 'doc_ticket_transferencia' in request.form else 0
    doc_ticket_otros = 1 if 'doc_ticket_otros' in request.form else 0
    doc_constancia_cbu = 1 if 'doc_constancia_cbu' in request.form else 0
    doc_constancia_cuit = 1 if 'doc_constancia_cuit' in request.form else 0
    doc_otros = 1 if 'doc_otros' in request.form else 0
    doc_otros_detalle = request.form['doc_otros_detalle']

    # Guardar formulario
    c.execute('''INSERT INTO formularios (
                 numero_formulario, fecha_emision, empresa_emite, tipo_orden,
                 origen_empresa, origen_cuenta, origen_detalle, origen_monto, origen_observacion, origen_mail, origen_whatsapp, origen_cargar, origen_tipo_destino,
                 destino_empresa, destino_cuenta, destino_detalle, destino_monto, destino_observacion, destino_mail, destino_whatsapp, destino_cargar, destino_tipo_destino,
                 doc_factura, doc_recibo, doc_fotocopia_cma, doc_presupuesto, doc_ticket_transferencia, doc_ticket_otros, doc_constancia_cbu, doc_constancia_cuit, doc_otros, doc_otros_detalle)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
              (numero_formulario, fecha_emision, empresa_emite, tipo_orden,
               origen_empresa, origen_cuenta, origen_detalle, origen_monto, origen_observacion, origen_mail, origen_whatsapp, origen_cargar, origen_tipo_destino,
               destino_empresa, destino_cuenta, destino_detalle, destino_monto, destino_observacion, destino_mail, destino_whatsapp, destino_cargar, destino_tipo_destino,
               doc_factura, doc_recibo, doc_fotocopia_cma, doc_presupuesto, doc_ticket_transferencia, doc_ticket_otros, doc_constancia_cbu, doc_constancia_cuit, doc_otros, doc_otros_detalle))
    formulario_id = c.lastrowid

    # Cuenta corriente
    for i in range(5):
        fecha = request.form.get(f'cc_fecha_{i}')
        detalle = request.form.get(f'cc_detalle_{i}')
        saldo_inicio = float(request.form.get(f'cc_saldo_inicio_{i}', 0))
        debe = float(request.form.get(f'cc_debe_{i}', 0))
        haber = float(request.form.get(f'cc_haber_{i}', 0))
        saldo_cierre = saldo_inicio + debe - haber
        if fecha or detalle or saldo_inicio or debe or haber:
            c.execute("INSERT INTO cuenta_corriente (formulario_id, fecha, detalle, saldo_inicio, debe, haber, saldo_cierre) VALUES (?, ?, ?, ?, ?, ?, ?)",
                      (formulario_id, fecha, detalle, saldo_inicio, debe, haber, saldo_cierre))

    # Forma de pago
    i = 0
    while f'pago_tipo_{i}' in request.form:
        tipo = request.form[f'pago_tipo_{i}']
        detalle = request.form[f'pago_detalle_{i}']
        monto = float(request.form[f'pago_monto_{i}'] or 0)
        c.execute("INSERT INTO forma_pago (formulario_id, tipo, detalle, monto) VALUES (?, ?, ?, ?)",
                  (formulario_id, tipo, detalle, monto))
        i += 1

    # Forma de cobro
    i = 0
    while f'cobro_tipo_{i}' in request.form:
        tipo = request.form[f'cobro_tipo_{i}']
        detalle = request.form[f'cobro_detalle_{i}']
        monto = float(request.form[f'cobro_monto_{i}'] or 0)
        c.execute("INSERT INTO forma_cobro (formulario_id, tipo, detalle, monto) VALUES (?, ?, ?, ?)",
                  (formulario_id, tipo, detalle, monto))
        i += 1

    conn.commit()
    conn.close()

    # Guardar en Excel
    save_to_excel(formulario_id)

    return redirect(url_for('index'))

if __name__ == '__main__':
    init_db()
    app.run(debug=True)